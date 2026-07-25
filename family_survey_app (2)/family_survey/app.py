"""
Family Survey Form - Flask Web Application
Based on Sompura Kendriya Parishad Samiti "Parivar Darpan" survey form.

Run with:  python app.py
"""

import os
import re
import sqlite3
from datetime import datetime

from flask import (
    Flask, render_template, request, redirect,
    url_for, flash, jsonify, send_file
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO

# --------------------------------------------------------------------------
# App configuration
# --------------------------------------------------------------------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "database", "family_survey.db")

app = Flask(__name__)
app.secret_key = "change-this-secret-key-in-production"

# --------------------------------------------------------------------------
# Database helpers
# --------------------------------------------------------------------------
def get_db_connection():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def init_db():
    os.makedirs(os.path.join(BASE_DIR, "database"), exist_ok=True)
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        CREATE TABLE IF NOT EXISTS families (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            family_no TEXT,
            gotra TEXT,
            village TEXT,
            survey_name TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS members (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            family_id INTEGER NOT NULL,
            sr_no INTEGER,
            member_no INTEGER,
            member_name TEXT,
            father_name TEXT,
            relation TEXT,
            dob_or_age TEXT,
            education TEXT,
            marital_status TEXT,
            mobile TEXT,
            occupation TEXT,
            aadhaar TEXT,
            current_address TEXT,
            marriage_date TEXT,
            annual_income TEXT,
            blood_group TEXT,
            FOREIGN KEY (family_id) REFERENCES families (id) ON DELETE CASCADE
        )
    """)

    conn.commit()
    conn.close()


# --------------------------------------------------------------------------
# Validation helpers
# --------------------------------------------------------------------------
def is_valid_aadhaar(value):
    if not value:
        return True  # optional field
    return bool(re.fullmatch(r"\d{12}", value.strip()))


def is_valid_mobile(value):
    if not value:
        return True  # optional field
    return bool(re.fullmatch(r"[6-9]\d{9}", value.strip()))


def is_valid_date(value):
    """Accepts empty, a plain age number, or a YYYY-MM-DD date string."""
    if not value:
        return True
    value = value.strip()
    if value.isdigit():
        # treated as age
        return 0 < int(value) < 120
    try:
        datetime.strptime(value, "%Y-%m-%d")
        return True
    except ValueError:
        return False


def validate_member(m):
    errors = []
    if not m.get("member_name", "").strip():
        errors.append(f"Member #{m.get('sr_no')}: नाम आवश्यक है (Name is required)")
    if not is_valid_aadhaar(m.get("aadhaar", "")):
        errors.append(f"Member #{m.get('sr_no')}: आधार संख्या 12 अंकों की होनी चाहिए (Aadhaar must be 12 digits)")
    if not is_valid_mobile(m.get("mobile", "")):
        errors.append(f"Member #{m.get('sr_no')}: मोबाइल नंबर अमान्य है (Invalid 10-digit mobile number)")
    if not is_valid_date(m.get("dob_or_age", "")):
        errors.append(f"Member #{m.get('sr_no')}: जन्म तिथि/उम्र अमान्य है (Invalid DOB or age)")
    if not is_valid_date(m.get("marriage_date", "")):
        errors.append(f"Member #{m.get('sr_no')}: विवाह तिथि अमान्य है (Invalid marriage date)")
    return errors


def extract_members_from_form(form):
    """Reads repeated member fields (arrays) from the submitted form."""
    names = form.getlist("member_name[]")
    count = len(names)

    def g(field, i):
        lst = form.getlist(f"{field}[]")
        return lst[i].strip() if i < len(lst) else ""

    members = []
    for i in range(count):
        members.append({
            "sr_no": i + 1,
            "member_no": g("member_no", i) or str(i + 1),
            "member_name": g("member_name", i),
            "father_name": g("father_name", i),
            "relation": g("relation", i),
            "dob_or_age": g("dob_or_age", i),
            "education": g("education", i),
            "marital_status": g("marital_status", i),
            "mobile": g("mobile", i),
            "occupation": g("occupation", i),
            "aadhaar": g("aadhaar", i),
            "current_address": g("current_address", i),
            "marriage_date": g("marriage_date", i),
            "annual_income": g("annual_income", i),
            "blood_group": g("blood_group", i),
        })
    return members


# --------------------------------------------------------------------------
# Public routes - Survey Form
# --------------------------------------------------------------------------
@app.route("/")
def index():
    return render_template("index.html")


@app.route("/submit", methods=["POST"])
def submit():
    family_no = request.form.get("family_no", "").strip()
    gotra = request.form.get("gotra", "").strip()
    village = request.form.get("village", "").strip()
    survey_name = request.form.get("survey_name", "").strip()

    members = extract_members_from_form(request.form)

    if not members:
        flash("कम से कम एक सदस्य जोड़ें (Please add at least one family member).", "danger")
        return redirect(url_for("index"))

    all_errors = []
    for m in members:
        all_errors.extend(validate_member(m))

    if all_errors:
        for e in all_errors:
            flash(e, "danger")
        return redirect(url_for("index"))

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO families (family_no, gotra, village, survey_name) VALUES (?, ?, ?, ?)",
        (family_no, gotra, village, survey_name),
    )
    family_id = cur.lastrowid

    for m in members:
        cur.execute("""
            INSERT INTO members (
                family_id, sr_no, member_no, member_name, father_name, relation,
                dob_or_age, education, marital_status, mobile, occupation,
                aadhaar, current_address, marriage_date, annual_income, blood_group
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            family_id, m["sr_no"], m["member_no"], m["member_name"], m["father_name"],
            m["relation"], m["dob_or_age"], m["education"], m["marital_status"],
            m["mobile"], m["occupation"], m["aadhaar"], m["current_address"],
            m["marriage_date"], m["annual_income"], m["blood_group"]
        ))

    conn.commit()
    conn.close()

    flash("परिवार सर्वे सफलतापूर्वक सबमिट हो गया! (Family survey submitted successfully!)", "success")
    return redirect(url_for("index"))


# --------------------------------------------------------------------------
# Admin Dashboard
# --------------------------------------------------------------------------
@app.route("/dashboard")
def dashboard():
    search = request.args.get("q", "").strip()

    conn = get_db_connection()
    query = """
        SELECT f.id AS family_id, f.family_no, f.gotra, f.village, f.survey_name,
               f.created_at, m.id AS member_id, m.sr_no, m.member_no, m.member_name,
               m.father_name, m.relation, m.dob_or_age, m.education, m.marital_status,
               m.mobile, m.occupation, m.aadhaar, m.current_address, m.marriage_date,
               m.annual_income, m.blood_group
        FROM families f
        JOIN members m ON m.family_id = f.id
    """
    params = ()
    if search:
        query += """
            WHERE f.family_no LIKE ? OR f.gotra LIKE ? OR f.village LIKE ?
               OR m.member_name LIKE ? OR m.father_name LIKE ? OR m.mobile LIKE ?
               OR m.aadhaar LIKE ?
        """
        like = f"%{search}%"
        params = (like, like, like, like, like, like, like)

    query += " ORDER BY f.id DESC, m.sr_no ASC"
    rows = conn.execute(query, params).fetchall()
    conn.close()

    return render_template("dashboard.html", rows=rows, search=search)


@app.route("/delete/<int:family_id>", methods=["POST"])
def delete_family(family_id):
    conn = get_db_connection()
    conn.execute("DELETE FROM members WHERE family_id = ?", (family_id,))
    conn.execute("DELETE FROM families WHERE id = ?", (family_id,))
    conn.commit()
    conn.close()
    flash("रिकॉर्ड हटा दिया गया (Record deleted).", "success")
    return redirect(url_for("dashboard"))


@app.route("/edit/<int:family_id>")
def edit_family(family_id):
    conn = get_db_connection()
    family = conn.execute("SELECT * FROM families WHERE id = ?", (family_id,)).fetchone()
    members = conn.execute(
        "SELECT * FROM members WHERE family_id = ? ORDER BY sr_no ASC", (family_id,)
    ).fetchall()
    conn.close()

    if family is None:
        flash("रिकॉर्ड नहीं मिला (Record not found).", "danger")
        return redirect(url_for("dashboard"))

    return render_template("edit.html", family=family, members=members)


@app.route("/update/<int:family_id>", methods=["POST"])
def update_family(family_id):
    family_no = request.form.get("family_no", "").strip()
    gotra = request.form.get("gotra", "").strip()
    village = request.form.get("village", "").strip()
    survey_name = request.form.get("survey_name", "").strip()

    members = extract_members_from_form(request.form)

    if not members:
        flash("कम से कम एक सदस्य जोड़ें (Please add at least one family member).", "danger")
        return redirect(url_for("edit_family", family_id=family_id))

    all_errors = []
    for m in members:
        all_errors.extend(validate_member(m))

    if all_errors:
        for e in all_errors:
            flash(e, "danger")
        return redirect(url_for("edit_family", family_id=family_id))

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        "UPDATE families SET family_no=?, gotra=?, village=?, survey_name=? WHERE id=?",
        (family_no, gotra, village, survey_name, family_id),
    )

    # Simplify update logic: remove old members, re-insert current set
    cur.execute("DELETE FROM members WHERE family_id = ?", (family_id,))
    for m in members:
        cur.execute("""
            INSERT INTO members (
                family_id, sr_no, member_no, member_name, father_name, relation,
                dob_or_age, education, marital_status, mobile, occupation,
                aadhaar, current_address, marriage_date, annual_income, blood_group
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            family_id, m["sr_no"], m["member_no"], m["member_name"], m["father_name"],
            m["relation"], m["dob_or_age"], m["education"], m["marital_status"],
            m["mobile"], m["occupation"], m["aadhaar"], m["current_address"],
            m["marriage_date"], m["annual_income"], m["blood_group"]
        ))

    conn.commit()
    conn.close()

    flash("रिकॉर्ड अपडेट हो गया (Record updated successfully).", "success")
    return redirect(url_for("dashboard"))


# --------------------------------------------------------------------------
# Export to Excel (.xlsx) using openpyxl - one row per family member
# --------------------------------------------------------------------------
@app.route("/export")
def export_excel():
    conn = get_db_connection()
    rows = conn.execute("""
        SELECT f.family_no, f.gotra, f.village, f.survey_name,
               m.sr_no, m.member_no, m.member_name, m.father_name, m.relation,
               m.dob_or_age, m.education, m.marital_status, m.mobile, m.occupation,
               m.aadhaar, m.current_address, m.marriage_date, m.annual_income, m.blood_group
        FROM families f
        JOIN members m ON m.family_id = f.id
        ORDER BY f.id ASC, m.sr_no ASC
    """).fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Family Survey"

    # Column order mirrors the original printed "Parivar Darpan" PDF table:
    # परिवार क्रमांक, गौत्र, क्रमांक, सदस्य क्रमांक, सदस्य के नाम, पिता का नाम, ...
    headers = [
        "Family No (परिवार क्रमांक)", "Gotra (गौत्र)", "Sr No (क्रमांक)",
        "Member No (सदस्य क्रमांक)", "Member Name (सदस्य के नाम)",
        "Father's Name (पिता का नाम)", "Relation with Head (मुखिया से सम्बन्ध)",
        "DOB / Age (जन्म तिथि/उम्र)", "Education (शैक्षणिक योग्यता)",
        "Marital Status (वैवाहिक स्थिति)", "Mobile Number (चल दूरभाष)",
        "Occupation (व्यवसाय)", "Aadhaar Number (आधार संख्या)",
        "Current Address (वर्तमान मुख्यावास पता)", "Marriage Date (विवाह तिथि)",
        "Annual Family Income (वार्षिक आय)", "Blood Group (रक्त समूह)",
        "Village/Town (गाँव/शहर)", "Survey Name (सर्वेक्षण का नाम)"
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # क्रमांक = a running serial number across the whole export, just like the
    # continuously-numbered "क्रमांक" column in the printed survey sheet.
    for global_sr_no, r in enumerate(rows, start=1):
        ws.append([
            r["family_no"], r["gotra"], global_sr_no, r["member_no"],
            r["member_name"], r["father_name"], r["relation"],
            r["dob_or_age"], r["education"], r["marital_status"], r["mobile"],
            r["occupation"], r["aadhaar"], r["current_address"], r["marriage_date"],
            r["annual_income"], r["blood_group"], r["village"], r["survey_name"]
        ])

    # Auto-fit column widths (approximate) based on the data actually written
    for col_idx in range(1, len(headers) + 1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        max_len = max(
            len(str(ws.cell(row=row_idx, column=col_idx).value or ""))
            for row_idx in range(1, ws.max_row + 1)
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"family_survey_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# --------------------------------------------------------------------------
# API helper used by JS for live Aadhaar/Mobile uniqueness (optional, simple)
# --------------------------------------------------------------------------
@app.route("/api/validate", methods=["POST"])
def api_validate():
    data = request.get_json(force=True)
    field = data.get("field")
    value = data.get("value", "")

    if field == "aadhaar":
        valid = is_valid_aadhaar(value)
    elif field == "mobile":
        valid = is_valid_mobile(value)
    elif field == "date":
        valid = is_valid_date(value)
    else:
        valid = True

    return jsonify({"valid": valid})


# --------------------------------------------------------------------------
if __name__ == "__main__":
    init_db()
    app.run(debug=True, host="0.0.0.0", port=5000)
